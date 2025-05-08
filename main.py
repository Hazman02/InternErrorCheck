from fastapi import FastAPI, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from io import BytesIO

app = FastAPI()

# === Static files ===
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/", response_class=FileResponse)
async def home():
    return FileResponse(STATIC_DIR / "index.html", media_type="text/html")

@app.get("/ping")
async def ping():
    return {"status": "ok"}


def parse_ping_report(wb_bytes: bytes, col: str = "A"):
    wb = load_workbook(filename=BytesIO(wb_bytes), data_only=True)
    errors = []
    for ws in wb.worksheets:
        lines = [(cell.row, str(cell.value).strip()) for cell in ws[col] if cell.value]
        i = 0
        while i < len(lines):
            row_idx, text = lines[i]
            if text.startswith("IP Address:"):
                ip = text.split(":", 1)[1].strip()
                if i+1 < len(lines) and lines[i+1][1].startswith("Status:"):
                    status = lines[i+1][1].split(":", 1)[1].strip()
                    detail = ""
                    if "failed" in status.lower() and i+2 < len(lines):
                        detail = lines[i+2][1]
                    if "failed" in status.lower():
                        errors.append({
                            "sheet": ws.title,
                            "row": row_idx,
                            "IP Address": ip,
                            "Status": status,
                            "Error Detail": detail
                        })
                i += 1
            i += 1
    return errors


def parse_command_report(wb_bytes: bytes, col: str = "A"):
    wb = load_workbook(filename=BytesIO(wb_bytes), data_only=True)
    errors = []

    for ws in wb.worksheets:
        current_iface = None
        current_imse = None
        current_service = None
        lines = [(cell.row, str(cell.value).strip()) for cell in ws[col] if cell.value]

        for idx, (row_idx, text) in enumerate(lines):
            if text.startswith("Interface:"):
                current_iface = text.split(":", 1)[1].strip()
            elif text.startswith("IMSE:"):
                current_imse = text.split(":", 1)[1].strip()
            elif text.startswith("Service:"):
                current_service = text.split(":", 1)[1].strip()

            if text.lower().startswith("title:"):
                title = text.split(":", 1)[1].strip()
                status = ""
                result = ""
                details = []
                j = idx + 1
                while j < len(lines) and not lines[j][1].lower().startswith("title:"):
                    _, nxt = lines[j]
                    low = nxt.lower()
                    if low.startswith("status:"):
                        status = nxt.split(":", 1)[1].strip()
                    elif low.startswith("result:"):
                        result = nxt.split(":", 1)[1].strip()
                    else:
                        details.append(nxt)
                    j += 1

                is_error = (
                    status.lower() == "failed" or
                    (status.lower() == "success" and "ping is failed" in result.lower())
                )

                if is_error:
                    errors.append({
                        "sheet": ws.title,
                        "row": row_idx,
                        "Interface": current_iface,
                        "IMSE": current_imse,
                        "Service": current_service,
                        "Title": title,
                        "Status": status,
                        "Result": result,
                        "Error Detail": " | ".join(details)
                    })

    return errors


def detect_report_type(wb_bytes: bytes):
    wb = load_workbook(filename=BytesIO(wb_bytes), data_only=True)
    first = wb.worksheets[0]
    for cell in first["A"][:20]:
        if cell.value and str(cell.value).strip().lower().startswith("title:"):
            return True
    return False


@app.post("/summarize/")
async def summarize_errors(
    pre_file:  UploadFile = File(None),
    post_file: UploadFile = File(None)
):
    pre_bytes  = await pre_file.read()  if pre_file  else None
    post_bytes = await post_file.read() if post_file else None
    pre_errors  = []
    post_errors = []

    # parse pre
    if pre_bytes:
        if detect_report_type(pre_bytes):
            pre_errors = parse_command_report(pre_bytes, col='A')
        else:
            pre_errors = parse_ping_report(pre_bytes, col='A')

    # parse post
    if post_bytes:
        if detect_report_type(post_bytes):
            post_errors = parse_command_report(post_bytes, col='A')
        else:
            post_errors = parse_ping_report(post_bytes, col='B')

    # cross-file ping logic
    if (
        pre_bytes and post_bytes
        and not detect_report_type(pre_bytes)
        and not detect_report_type(post_bytes)
    ):
        pre_ips  = {e["IP Address"] for e in pre_errors}
        post_ips = {e["IP Address"] for e in post_errors}
        pre_map  = {e["IP Address"]:(e["sheet"], e["row"], e["Status"]) for e in pre_errors}
        for ip in pre_ips - post_ips:
            sheet_p, row_p, status_p = pre_map.get(ip, (None, None, None))
            post_errors.append({
                "sheet": sheet_p,
                "row":   row_p,
                "IP Address": ip,
                "Status": status_p,
                "Error Detail": "Pre-migration failure (Pre-migration Side)"
            })

    return JSONResponse({"pre_errors": pre_errors, "post_errors": post_errors})


def build_workbook_with_summary(wb_bytes: bytes, errors: list):
    wb = load_workbook(filename=BytesIO(wb_bytes))
    ws = wb.create_sheet(title="Error Summary")

    if not errors:
        ws["A1"] = "No errors found."
    else:
        # 1) Headers
        headers     = list(errors[0].keys())
        header_fill = PatternFill(patternType="solid", fgColor="7B1FA2")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2) Data rows with alternating fills
        even_fill = PatternFill(patternType="solid", fgColor="DDDDDD")
        odd_fill  = PatternFill(patternType="solid", fgColor="FFFFFF")

        for r, err in enumerate(errors, start=2):
            fill = even_fill if (r % 2 == 0) else odd_fill
            for c, key in enumerate(headers, start=1):
                v    = err.get(key, "")
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill
                cell.alignment = Alignment(
                    horizontal="left" if c > 1 else "center",
                    vertical="center",
                    wrap_text=True,
                    indent=1
                )

        # 3) Auto-size columns (up to a max width)
        for idx in range(1, len(headers) + 1):
            col = get_column_letter(idx)
            max_length = max(
                len(str(ws[f"{col}{row}"].value or "")) for row in range(1, len(errors) + 2)
            )
            ws.column_dimensions[col].width = min(max_length + 4, 50)

        # 4) Freeze header and enable filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(errors)+1}"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

@app.post("/download-pre/")
async def download_pre(pre_file: UploadFile = File(...)):
    data = await pre_file.read()
    if detect_report_type(data):
        errors = parse_command_report(data, col="A")
    else:
        errors = parse_ping_report(data, col="A")
    out = build_workbook_with_summary(data, errors)
    fn = pre_file.filename.replace(".xlsx","_with_summary.xlsx")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


@app.post("/download-post/")
async def download_post(post_file: UploadFile = File(...)):
    data = await post_file.read()
    if detect_report_type(data):
        errors = parse_command_report(data, col="A")
    else:
        errors = parse_ping_report(data, col="B")
    out = build_workbook_with_summary(data, errors)
    fn = post_file.filename.replace(".xlsx","_with_summary.xlsx")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )
